import numpy as np
import math
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series

# import matplotlib.pyplot as plt


def st(n):  # convert column number into excel ABC format
    text = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"  # 26 base
    if n == 0:
        return "A"
    base = 26
    t = ""
    while n > 0:
        n1 = n % base
        t = text[n1 : n1 + 1] + t
        n = int(n / base)
    return t


def sts(i, j):  # convert column-row
    t = st(i) + str(j + 1)
    return t


class Pigment:
    def __init__(self):
        self.coord = np.zeros((3))
        self.mu = np.zeros((3))
        self.mag = 0.0


"""
------------------------------------------------------------
Using class Exciton
- Constructor takes two filenames
  The first is xlsx (excel) input file name, see Dimer.xlsx for structure
  only 3 pages are important, General, Hamiltonian and Pigment
  All numbers must be exactly in cells where they are, the program
  reads them by cell number and sheet number.
  The second file is output file. It can be the same as input, but that file
  must not be open before starting the program. It will keep 3 first pages intact
  and will append to them the results of calculation if second file is omitted,
  then it will run faster and will not create output file. All elements can be
  accessed by class members
- Once constructed, the class diagonalizes matrix and calculates spectra, they
  can be accesses as following members of class:
  filename - (Exciton.filename) input file name
  wb - excel workbook with input parameters, and output as well if output file was defined
  size - the size of the system (number of pigments, or hamiltonian sizexsize)
  bandwidth - the bandwidth to broaden sticks with (from excel input)
              if bandwidth was zero, no spectra will be calculated!
  xfrom, xto, xstep - limits and step for spectra calculation range
  ham - hamiltonian
  pig[i] - array with pigment Pigment structures (size elements), from input file
       pig.coord - vector showing the position of pigment
       pig.mu - transition dipole moment vector
       pig.mag - the magnitude of dipole. If it was 0 in input file, then
                 input mu will be as in the file, otherwise vector mu
                 is rescaled to have magnitude mag, but original direction
  eval[i] - eigenvalues obtained by dioganalizing input ham
  evec[j,i] - eigenvectors corresponding to eval[i]
  mu[i] - exciton transition dipole moment vectors
  stickA[i] - absorption strength of exciton band (mu[i]**2)
  stickCD[i] - rotational strength of exciton band
  in case xstep was not zero, spectra are calculated:
  x[i] - x-axis for xpectra
  abs[i] - absorption spectrum
  CD[i] - CD spectrum
"""


class Exciton:
    def __init__(self, filename):
        self.filename = filename
        self.wb = wb = load_workbook(self.filename)
        sheet = wb.sheetnames
        # Now delete all output worksheets above #3, i.e. keep only 3 poages!!!
        for i in range(3, len(sheet)):
            self.wb.remove(wb[sheet[i]])
        self.size = N = wb[sheet[0]]["B10"].value
        self.bandwidth = wb[sheet[0]]["B11"].value
        self.xfrom = wb[sheet[0]]["B12"].value
        self.xto = wb[sheet[0]]["C12"].value
        self.xstep = wb[sheet[0]]["D12"].value
        self.ham = np.zeros((N, N))
        self.origham = np.zeros((N, N))
        for i in range(0, N):
            for j in range(0, N):
                self.ham[i][j] = self.origham[i, j] = wb[sheet[1]][i + 1][j].value
        # self.origham = self.ham;
        self.pig = np.ndarray((N), Pigment)
        self.origpig = np.ndarray((N), Pigment)
        for i in range(0, N):
            self.pig[i] = Pigment()
            self.pig[i].mag = wb[sheet[2]][sts(1, i + 2)].value
            for j in range(0, 3):
                self.pig[i].mu[j] = wb[sheet[2]][sts(j + 2, i + 2)].value
                self.pig[i].coord[j] = wb[sheet[2]][sts(j + 5, i + 2)].value
        # normalize mu to mag
        if self.pig[i].mag != 0:
            for i in range(0, N):
                magn = np.dot(self.pig[i].mu, self.pig[i].mu)
                self.pig[i].mu = self.pig[i].mu * (self.pig[i].mag / magn)
        self.origpig = self.pig
        # diagonalize Hamiltonian
        # self.eval, self.evec =  np.linalg.eig(self.ham)
        if self.xstep != 0:
            self.x = np.arange(self.xfrom, self.xto, self.xstep)
            self.abs = self.x * 0.0
            self.CD = self.x * 0.0
        # self.recalculate()
        # write to workbook: only if output book name is defined!
        return

    def diagonalize(self):
        self.eval, self.evec = np.linalg.eigh(self.ham)
        return

    def getsticks(self):
        N = self.size
        self.mu = np.zeros((N, 3))
        self.stickA = np.zeros((N))
        self.stickCD = np.zeros((N))
        for i in range(0, N):
            for j in range(0, N):
                self.mu[i] += self.evec[j, i] * self.pig[j].mu
            self.stickA[i] = np.dot(self.mu[i], self.mu[i])
        # CD stick spectrum
        for i in range(0, N):
            for j in range(0, N):
                for k in range(0, N):
                    dist = self.pig[k].coord - self.pig[j].coord
                    mu_cross = np.cross(self.pig[j].mu, self.pig[k].mu)
                    self.stickCD[i] += (
                        self.evec[j, i] * self.evec[k, i] * np.dot(dist, mu_cross)
                    )
        return

    def getspectra(self):
        if self.xstep == 0:
            return
        N = self.size
        if self.xstep != 0:
            sigma2 = self.bandwidth ** 2 / (4.0 * math.log(2.0))
            for i in range(0, N):
                self.abs += self.stickA[i] * np.exp(
                    -(self.x - self.eval[i]) ** 2 / sigma2
                )
                self.CD += self.stickCD[i] * np.exp(
                    -(self.x - self.eval[i]) ** 2 / sigma2
                )
        return

    def recalculate(self):
        self.diagonalize()
        self.getsticks()
        self.getspectra()
        return

    def save_excel(self, filename):
        sheet = self.wb.sheetnames
        # Now delete all output worksheets above #3, i.e. keep only 3 poages!!!
        for i in range(3, len(sheet)):
            self.wb.remove(wb[sheet[i]])
        N = self.size
        ws_eig = self.wb.create_sheet()
        ws_eig.title = "Sticks"
        ws_eig[sts(0, 0)].value = "Eval"
        ws_eig[sts(1, 0)].value = "Abs"
        ws_eig[sts(2, 0)].value = "Rot"
        ws_eig[sts(3, 0)].value = "Evector"
        for i in range(0, N):
            ws_eig[sts(0, i + 1)].value = self.eval[i]
            ws_eig[sts(1, i + 1)].value = self.stickA[i]
            ws_eig[sts(2, i + 1)].value = self.stickCD[i]
            for j in range(0, N):
                ws_eig[sts(3 + j, i + 1)].value = self.evec[j, i]
                ws_spec = self.wb.create_sheet()
                ws_spec.title = "Spectra"
        for i in range(0, np.size(self.x)):
            ws_spec[sts(0, i)].value = self.x[i]
            ws_spec[sts(1, i)].value = self.abs[i]
            ws_spec[sts(2, i)].value = self.CD[i]
        # add chart
        c1 = ScatterChart()
        c1.title = ""
        c1.x_axis.title = "wavenumber"
        c1.y_axis.title = "Absorbance"
        c1.style = 13
        xvalues = Reference(ws_spec, min_col=1, min_row=1, max_row=len(self.x))
        values = Reference(
            ws_spec, min_col=2, max_col=2, min_row=1, max_row=len(self.x)
        )
        series = Series(values, xvalues)
        series.graphicalProperties.line.solidFill = "FF0000"
        series.graphicalProperties.line.width = 3
        c1.series.append(series)
        c1.x_axis.scaling.min = self.x[0]
        c1.x_axis.scaling.max = self.x[len(self.x) - 1]
        c1.x_axis.tickLblPos = "low"
        c1.y_axis.tickLblPos = "low"
        c1.legend = None
        ws_spec.add_chart(c1, "D2")
        c2 = ScatterChart()
        c2.title = ""
        c2.x_axis.title = "wavenumber"
        c2.y_axis.title = "CD"
        c2.style = 13
        # xvalues = Reference(ws_spec, min_col=1, min_row=1, max_row=len(self.x))
        values = Reference(
            ws_spec, min_col=3, max_col=3, min_row=1, max_row=len(self.x)
        )
        series = Series(values, xvalues)
        series.graphicalProperties.line.solidFill = "0000FF"
        series.graphicalProperties.line.width = 3
        c2.series.append(series)
        c2.x_axis.scaling.min = self.x[0]
        c2.x_axis.scaling.max = self.x[len(self.x) - 1]
        c2.x_axis.tickLblPos = "low"
        c2.y_axis.tickLblPos = "low"
        c2.legend = None
        ws_spec.add_chart(c2, "D17")
        self.wb.save(filename)
        return
